import openpyxl
import json

# ── Coordenades aproximades per municipi ─────────────────
COORDS = {
    "ALCOVER": (41.2769, 1.1774),
    "VALLS": (41.2867, 1.2500),
    "TARRAGONA": (41.1189, 1.2445),
    "REUS": (41.1561, 1.1065),
    "CAMBRILS": (41.0674, 1.0580),
    "SALOU": (41.0752, 1.1422),
    "VILA-SECA": (41.1047, 1.1469),
    "TORREDEMBARRA": (41.1467, 1.3961),
    "EL VENDRELL": (41.2183, 1.5358),
    "CALAFELL": (41.2000, 1.5667),
    "AMPOSTA": (40.7150, 0.5800),
    "TORTOSA": (40.8122, 0.5211),
    "LA SEU D'URGELL": (42.3583, 1.4567),
    "MONTBLANC": (41.3833, 1.1667),
    "FALSET": (41.1500, 0.8167),
}

def get_coords(municipi):
    if not municipi:
        return (41.1189, 1.2445)
    for key, coords in COORDS.items():
        if key in str(municipi).upper():
            return coords
    return (41.1189, 1.2445)

wb = openpyxl.load_workbook("C:/ELOI/PROGRAMACION/PROYECTOS/HACKATO_IA/adreces_i_telefons.xlsx", read_only=True)

# ── 1. CENTRES ────────────────────────────────────────────
centres = []
ws1 = wb["CÀRITAS ARXIPRESTAT"]
rows = list(ws1.iter_rows(values_only=True))
current_arxiprestat = None
cid = 1

for row in rows:
    if not any(row):
        continue
    if row[0] and not row[1]:
        current_arxiprestat = row[0]
        continue
    if row[0] == "CÀRITAS":
        continue
    if row[0] and row[1]:
        lat, lng = get_coords(row[3] or row[0])
        centres.append({
            "id": f"CTR{cid:03d}",
            "nom": str(row[0]),
            "arxiprestat": current_arxiprestat,
            "adreça": str(row[1]) if row[1] else "",
            "codi_postal": str(row[2]) if row[2] else "",
            "municipi": str(row[3]) if row[3] else "",
            "email": str(row[4]) if row[4] else "",
            "horari": str(row[5]) if row[5] else "",
            "lat": lat,
            "lng": lng
        })
        cid += 1

with open("centres.json", "w", encoding="utf-8") as f:
    json.dump(centres, f, ensure_ascii=False, indent=2)
print(f"✅ centres.json: {len(centres)} centres")

# ── 2. VOLUNTARIS (sintètics basats en Sheet 2) ───────────
ws2 = wb["PERSONES I PROJECTES"]
rows2 = list(ws2.iter_rows(values_only=True))
voluntaris = []
vid = 1

for row in rows2:
    if not row[4]:
        continue
    if row[4] in ["PERSONES /PROJECTES", None]:
        continue
    if row[3] and str(row[3]) in ["Direcció", "Administració", "Coordinació"]:
        municipi = str(row[0]).split("/")[0].strip() if row[0] else "Tarragona"
        lat, lng = get_coords(municipi)
        voluntaris.append({
            "id": f"VOL{vid:03d}",
            "nom": str(row[4]),
            "rol": str(row[3]),
            "projecte": str(row[2]) if row[2] else "General",
            "email": str(row[5]) if row[5] else "",
            "telefon": str(row[7]) if row[7] else "",
            "municipi": municipi,
            "lat": lat,
            "lng": lng,
            "habilitats": ["acompanyament", "orientació", "gestió"],
            "disponibilitat": ["dilluns_mati", "dimecres_tarda"],
            "max_persones": 5,
            "persones_actuals": 0
        })
        vid += 1

# Afegim voluntaris sintètics addicionals
VOLUNTARIS_SINT = [
    {"habilitats": ["alimentació", "logística", "repartiment"], "disponibilitat": ["dimarts_tarda", "dijous_tarda"], "municipi": "Tarragona"},
    {"habilitats": ["assessoria_jurídica", "situació_irregular", "documentació"], "disponibilitat": ["dilluns_tarda", "divendres_mati"], "municipi": "Reus"},
    {"habilitats": ["inserció_laboral", "formació", "orientació_professional"], "disponibilitat": ["dimecres_mati", "dijous_mati"], "municipi": "Valls"},
    {"habilitats": ["atenció_infants", "suport_escolar", "família"], "disponibilitat": ["dilluns_tarda", "dimarts_tarda"], "municipi": "Tarragona"},
    {"habilitats": ["salut_mental", "acompanyament_psicològic", "gent_gran"], "disponibilitat": ["dimecres_tarda", "divendres_tarda"], "municipi": "Cambrils"},
    {"habilitats": ["habitatge", "mediació", "desnonament"], "disponibilitat": ["dimarts_mati", "dijous_mati"], "municipi": "Tarragona"},
    {"habilitats": ["català", "castellà", "àrab", "traducció", "acollida"], "disponibilitat": ["dilluns_mati", "dimecres_mati"], "municipi": "Reus"},
]
NOMS = ["Anna P.", "Pere M.", "Laura F.", "Jordi S.", "Montse R.", "Carles B.", "Fàtima A."]
for i, v in enumerate(VOLUNTARIS_SINT):
    lat, lng = get_coords(v["municipi"])
    voluntaris.append({
        "id": f"VOL{vid:03d}",
        "nom": NOMS[i],
        "rol": "Voluntari/a",
        "projecte": "Acollida i Acompanyament",
        "email": f"voluntari{vid}@caritasdtarragona.cat",
        "telefon": "",
        "municipi": v["municipi"],
        "lat": lat,
        "lng": lng,
        "habilitats": v["habilitats"],
        "disponibilitat": v["disponibilitat"],
        "max_persones": 4,
        "persones_actuals": 0
    })
    vid += 1

with open("voluntaris.json", "w", encoding="utf-8") as f:
    json.dump(voluntaris, f, ensure_ascii=False, indent=2)
print(f"✅ voluntaris.json: {len(voluntaris)} voluntaris")

# ── 3. RECURSOS (sintètics) ───────────────────────────────
recursos = [
    {"id":"REC001","nom":"Ajuda alimentària setmanal","tipus":"alimentació",
     "keywords":["alimentació","fam","sense_ingressos","família","infants","pobresa"],
     "quantitat_disponible":50,"unitat":"cistelles/setmana","centre_id":"CTR001",
     "requisits":["empadronat","ingressos_baixos"]},
    {"id":"REC002","nom":"Ajuda pagament lloguer","tipus":"habitatge",
     "keywords":["habitatge","lloguer","desnonament","inestabilitat_habitacional"],
     "quantitat_disponible":10,"unitat":"ajudes/mes","centre_id":"CTR001",
     "requisits":["ingressos_baixos","risc_exclusio"]},
    {"id":"REC003","nom":"Roba i equipament llar","tipus":"espècie",
     "keywords":["roba","equipament","necessitats_bàsiques","família","infants"],
     "quantitat_disponible":100,"unitat":"lots","centre_id":"CTR001","requisits":[]},
    {"id":"REC004","nom":"Assessoria jurídica immigració","tipus":"legal",
     "keywords":["situació_irregular","documentació","immigració","extracomunitari"],
     "quantitat_disponible":20,"unitat":"sessions/mes","centre_id":"CTR001",
     "requisits":["situació_irregular"]},
    {"id":"REC005","nom":"Orientació laboral i CV","tipus":"inserció_laboral",
     "keywords":["atur","sense_feina","inserció_laboral","formació","joves"],
     "quantitat_disponible":30,"unitat":"sessions/mes","centre_id":"CTR001","requisits":[]},
    {"id":"REC006","nom":"Suport escolar infants","tipus":"educació",
     "keywords":["infants","educació","suport_escolar","família","menors"],
     "quantitat_disponible":25,"unitat":"places","centre_id":"CTR001","requisits":["menors_a_carrec"]},
    {"id":"REC007","nom":"Atenció psicosocial","tipus":"salut_mental",
     "keywords":["salut_mental","ansietat","acompanyament_psicològic","soledat","gent_gran"],
     "quantitat_disponible":15,"unitat":"sessions/mes","centre_id":"CTR001","requisits":[]},
    {"id":"REC008","nom":"Transport i acompanyament sanitari","tipus":"salut",
     "keywords":["salut","malaltia","mobilitat","gent_gran","discapacitat"],
     "quantitat_disponible":20,"unitat":"serveis/mes","centre_id":"CTR001","requisits":[]},
]
with open("recursos.json", "w", encoding="utf-8") as f:
    json.dump(recursos, f, ensure_ascii=False, indent=2)
print(f"✅ recursos.json: {len(recursos)} recursos")

# ── 4. ORGANITZACIONS (sintètiques) ──────────────────────
organitzacions = [
    {"id":"ORG001","nom":"Creu Roja Tarragona","tipus":"ONG",
     "serveis":["alimentació","roba","acompanyament_sanitari","transport"],
     "keywords":["salut","malaltia","mobilitat","gent_gran","urgència"],
     "contacte":"tarragona@creuroja.org","municipi":"Tarragona",
     "lat":41.121,"lng":1.245,"conveni_caritas":True},
    {"id":"ORG002","nom":"SAIER Tarragona","tipus":"servei_públic",
     "serveis":["assessoria_jurídica","documentació","acollida_immigrants"],
     "keywords":["situació_irregular","immigració","documentació","extracomunitari"],
     "contacte":"saier@tarragona.cat","municipi":"Tarragona",
     "lat":41.119,"lng":1.243,"conveni_caritas":True},
    {"id":"ORG003","nom":"Hàbitat3 Tarragona","tipus":"ONG",
     "serveis":["habitatge_social","mediació_habitatge","allotjament_temporal"],
     "keywords":["habitatge","desnonament","sense_llar","urgència_habitatge"],
     "contacte":"tarragona@habitat3.org","municipi":"Tarragona",
     "lat":41.118,"lng":1.246,"conveni_caritas":True},
    {"id":"ORG004","nom":"SOC Reus","tipus":"servei_públic",
     "serveis":["orientació_laboral","formació_ocupacional","borsa_treball"],
     "keywords":["atur","inserció_laboral","formació","sense_feina"],
     "contacte":"soc.reus@gencat.cat","municipi":"Reus",
     "lat":41.156,"lng":1.106,"conveni_caritas":False},
]
with open("organitzacions.json", "w", encoding="utf-8") as f:
    json.dump(organitzacions, f, ensure_ascii=False, indent=2)
print(f"✅ organitzacions.json: {len(organitzacions)} organitzacions")

# ── 5. EMPRESES (sintètiques - Empreses amb Cor) ─────────
empreses = [
    {"id":"EMP001","nom":"Repsol Tarragona","tipus_col·laboració":["voluntariat_corporatiu","donació_econòmica"],
     "recursos_oferts":["formació_laboral","orientació_professional"],
     "keywords":["inserció_laboral","formació","joves","atur"],
     "contacte":"rse@repsol.com","hores_voluntariat_disponibles":40},
    {"id":"EMP002","nom":"Port de Tarragona","tipus_col·laboració":["donació_espècie","voluntariat_corporatiu"],
     "recursos_oferts":["material_oficina","espais_formació"],
     "keywords":["formació","educació","joves"],
     "contacte":"rse@porttarragona.cat","hores_voluntariat_disponibles":20},
    {"id":"EMP003","nom":"Mercadona Tarragona","tipus_col·laboració":["donació_espècie"],
     "recursos_oferts":["alimentació","productes_primera_necessitat"],
     "keywords":["alimentació","necessitats_bàsiques","família","pobresa"],
     "contacte":"donacions@mercadona.es","hores_voluntariat_disponibles":0},
]
with open("empreses.json", "w", encoding="utf-8") as f:
    json.dump(empreses, f, ensure_ascii=False, indent=2)
print(f"✅ empreses.json: {len(empreses)} empreses")

print("\n🎉 Totes les BBDDs generades correctament!")
